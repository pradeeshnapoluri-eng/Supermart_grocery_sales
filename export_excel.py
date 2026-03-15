import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings("ignore")

df          = pd.read_csv("grocery_sales_cleaned.csv")
cat_stats   = pd.read_csv("sql_category_stats.csv")
reg_stats   = pd.read_csv("sql_region_stats.csv")
yearly      = pd.read_csv("sql_yearly_stats.csv")
monthly     = pd.read_csv("sql_monthly_stats.csv")
top_cities  = pd.read_csv("sql_top_cities.csv")
loss_orders = pd.read_csv("sql_loss_orders.csv")
top_cust    = pd.read_csv("sql_top_customers.csv")
profit_rate = pd.read_csv("sql_profit_rate.csv")
ml_results  = pd.read_csv("ml_model_results.csv")
overall     = pd.read_csv("sql_overall_stats.csv")

print("All CSV files loaded!")

wb = Workbook()

header_font    = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
header_fill    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
sub_font       = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
sub_fill       = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
alt_fill       = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
title_font     = Font(name="Calibri", bold=True, size=16, color="1F4E79")
center_align   = Alignment(horizontal="center", vertical="center")
left_align     = Alignment(horizontal="left",   vertical="center")
green_fill     = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill       = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
gold_fill      = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

thin_border = Border(
    left   = Side(style="thin"),
    right  = Side(style="thin"),
    top    = Side(style="thin"),
    bottom = Side(style="thin")
)

def style_header(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell            = ws.cell(row=row_num, column=col)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = center_align
        cell.border     = thin_border

def style_row(ws, row_num, num_cols, alternate=False):
    for col in range(1, num_cols + 1):
        cell           = ws.cell(row=row_num, column=col)
        if alternate:
            cell.fill  = alt_fill
        cell.alignment = center_align
        cell.border    = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 4

print("Creating Sheet 1 - Overview...")

ws1 = wb.active
ws1.title = "Overview"

ws1.merge_cells("A1:H1")
ws1["A1"]           = "SUPERMART GROCERY SALES - RETAIL ANALYTICS REPORT"
ws1["A1"].font      = title_font
ws1["A1"].alignment = center_align
ws1.row_dimensions[1].height = 40

ws1.append([])

ws1.append(["Total Orders", "Total Sales", "Total Profit",
            "Avg Sales", "Avg Profit", "Avg Discount",
            "Min Sales", "Max Sales"])
style_header(ws1, 3, 8)

row_data = [
    int(overall["total_orders"].iloc[0]),
    float(overall["total_sales"].iloc[0]),
    float(overall["total_profit"].iloc[0]),
    float(overall["avg_sales"].iloc[0]),
    float(overall["avg_profit"].iloc[0]),
    float(overall["avg_discount"].iloc[0]),
    float(overall["min_sales"].iloc[0]),
    float(overall["max_sales"].iloc[0])
]
ws1.append(row_data)
style_row(ws1, 4, 8)

ws1.append([])
ws1.append(["Category Performance Summary"])
ws1.cell(row=6, column=1).font = Font(bold=True, size=13, color="1F4E79")

cat_headers = ["Category", "Total Orders", "Total Sales",
               "Total Profit", "Avg Sales", "Avg Profit"]
ws1.append(cat_headers)
style_header(ws1, 7, len(cat_headers))

for i, row in enumerate(cat_stats.itertuples(index=False)):
    ws1.append(list(row))
    row_num  = 8 + i
    style_row(ws1, row_num, len(cat_headers), alternate=(i % 2 == 0))
    profit_cell = ws1.cell(row=row_num, column=4)
    try:
        val = float(profit_cell.value)
        if val > 0:
            profit_cell.fill = green_fill
        else:
            profit_cell.fill = red_fill
    except:
        pass

auto_width(ws1)
print("Sheet 1 done!")

print("Creating Sheet 2 - Raw Data...")

ws2 = wb.create_sheet("Raw Data")

raw_cols = ["Order_ID", "Customer_Name", "Category", "Sub_Category",
            "City", "Order_Date", "Region", "Sales", "Discount",
            "Profit", "State", "month_no", "Month", "year",
            "Profit_Margin", "Discount_Amount"]

df_raw = df.copy()
df_raw.columns = df_raw.columns.str.replace(" ", "_")

available = [c for c in raw_cols if c in df_raw.columns]
df_export = df_raw[available].copy()

ws2.append(available)
style_header(ws2, 1, len(available))

for i, row in enumerate(df_export.itertuples(index=False)):
    ws2.append(list(row))
    style_row(ws2, i + 2, len(available), alternate=(i % 2 == 0))

auto_width(ws2)
print("Sheet 2 done!")

print("Creating Sheet 3 - Category Analysis...")

ws3 = wb.create_sheet("Category Analysis")

ws3.merge_cells("A1:F1")
ws3["A1"]           = "Sales and Profit Analysis by Category"
ws3["A1"].font      = title_font
ws3["A1"].alignment = center_align
ws3.row_dimensions[1].height = 30

ws3.append(list(cat_stats.columns))
style_header(ws3, 2, len(cat_stats.columns))

for i, row in enumerate(cat_stats.itertuples(index=False)):
    ws3.append(list(row))
    row_num = i + 3
    style_row(ws3, row_num, len(cat_stats.columns), alternate=(i % 2 == 0))

chart1 = BarChart()
chart1.title       = "Total Sales by Category"
chart1.y_axis.title = "Sales"
chart1.x_axis.title = "Category"
chart1.width        = 22
chart1.height       = 14

data_ref = Reference(ws3, min_col=3, min_row=2, max_row=len(cat_stats) + 2)
cats_ref = Reference(ws3, min_col=1, min_row=3, max_row=len(cat_stats) + 2)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
ws3.add_chart(chart1, "H2")

auto_width(ws3)
print("Sheet 3 done!")

print("Creating Sheet 4 - Region Analysis...")

ws4 = wb.create_sheet("Region Analysis")

ws4.merge_cells("A1:E1")
ws4["A1"]           = "Sales and Profit Analysis by Region"
ws4["A1"].font      = title_font
ws4["A1"].alignment = center_align
ws4.row_dimensions[1].height = 30

ws4.append(list(reg_stats.columns))
style_header(ws4, 2, len(reg_stats.columns))

for i, row in enumerate(reg_stats.itertuples(index=False)):
    ws4.append(list(row))
    style_row(ws4, i + 3, len(reg_stats.columns), alternate=(i % 2 == 0))

chart2 = PieChart()
chart2.title  = "Sales Distribution by Region"
chart2.width  = 18
chart2.height = 14

data_ref2 = Reference(ws4, min_col=3, min_row=2, max_row=len(reg_stats) + 2)
cats_ref2 = Reference(ws4, min_col=1, min_row=3, max_row=len(reg_stats) + 2)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)
ws4.add_chart(chart2, "H2")

auto_width(ws4)
print("Sheet 4 done!")

print("Creating Sheet 5 - Yearly Analysis...")

ws5 = wb.create_sheet("Yearly Analysis")

ws5.merge_cells("A1:E1")
ws5["A1"]           = "Yearly Sales and Profit Trend"
ws5["A1"].font      = title_font
ws5["A1"].alignment = center_align
ws5.row_dimensions[1].height = 30

ws5.append(list(yearly.columns))
style_header(ws5, 2, len(yearly.columns))

for i, row in enumerate(yearly.itertuples(index=False)):
    ws5.append(list(row))
    style_row(ws5, i + 3, len(yearly.columns), alternate=(i % 2 == 0))

chart3 = LineChart()
chart3.title        = "Yearly Sales Trend"
chart3.y_axis.title = "Sales"
chart3.x_axis.title = "Year"
chart3.width        = 22
chart3.height       = 14

data_ref3 = Reference(ws5, min_col=3, min_row=2, max_row=len(yearly) + 2)
cats_ref3 = Reference(ws5, min_col=1, min_row=3, max_row=len(yearly) + 2)
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref3)
ws5.add_chart(chart3, "H2")

auto_width(ws5)
print("Sheet 5 done!")

print("Creating Sheet 6 - Monthly Analysis...")

ws6 = wb.create_sheet("Monthly Analysis")

ws6.merge_cells("A1:F1")
ws6["A1"]           = "Monthly Sales and Profit Trend"
ws6["A1"].font      = title_font
ws6["A1"].alignment = center_align
ws6.row_dimensions[1].height = 30

ws6.append(list(monthly.columns))
style_header(ws6, 2, len(monthly.columns))

for i, row in enumerate(monthly.itertuples(index=False)):
    ws6.append(list(row))
    style_row(ws6, i + 3, len(monthly.columns), alternate=(i % 2 == 0))

chart4 = LineChart()
chart4.title        = "Monthly Sales Trend"
chart4.y_axis.title = "Sales"
chart4.x_axis.title = "Month"
chart4.width        = 22
chart4.height       = 14

data_ref4 = Reference(ws6, min_col=4, min_row=2, max_row=len(monthly) + 2)
cats_ref4 = Reference(ws6, min_col=2, min_row=3, max_row=len(monthly) + 2)
chart4.add_data(data_ref4, titles_from_data=True)
chart4.set_categories(cats_ref4)
ws6.add_chart(chart4, "H2")

auto_width(ws6)
print("Sheet 6 done!")

print("Creating Sheet 7 - Top Cities...")

ws7 = wb.create_sheet("Top Cities")

ws7.merge_cells("A1:D1")
ws7["A1"]           = "Top 10 Cities by Sales"
ws7["A1"].font      = title_font
ws7["A1"].alignment = center_align
ws7.row_dimensions[1].height = 30

ws7.append(list(top_cities.columns))
style_header(ws7, 2, len(top_cities.columns))

for i, row in enumerate(top_cities.itertuples(index=False)):
    ws7.append(list(row))
    style_row(ws7, i + 3, len(top_cities.columns), alternate=(i % 2 == 0))

chart5 = BarChart()
chart5.title        = "Top 10 Cities by Sales"
chart5.y_axis.title = "Sales"
chart5.x_axis.title = "City"
chart5.width        = 22
chart5.height       = 14

data_ref5 = Reference(ws7, min_col=2, min_row=2, max_row=len(top_cities) + 2)
cats_ref5 = Reference(ws7, min_col=1, min_row=3, max_row=len(top_cities) + 2)
chart5.add_data(data_ref5, titles_from_data=True)
chart5.set_categories(cats_ref5)
ws7.add_chart(chart5, "F2")

auto_width(ws7)
print("Sheet 7 done!")

print("Creating Sheet 8 - Loss Orders...")

ws8 = wb.create_sheet("Loss Orders")

ws8.merge_cells("A1:F1")
ws8["A1"]           = "Loss Making Orders (Negative Profit)"
ws8["A1"].font      = Font(name="Calibri", bold=True, size=14, color="9C0006")
ws8["A1"].alignment = center_align
ws8.row_dimensions[1].height = 30

ws8.append(list(loss_orders.columns))
style_header(ws8, 2, len(loss_orders.columns))

for i, row in enumerate(loss_orders.itertuples(index=False)):
    ws8.append(list(row))
    row_num = i + 3
    style_row(ws8, row_num, len(loss_orders.columns), alternate=(i % 2 == 0))
    profit_cell      = ws8.cell(row=row_num, column=6)
    profit_cell.fill = red_fill
    profit_cell.font = Font(bold=True, color="9C0006")

auto_width(ws8)
print("Sheet 8 done!")

print("Creating Sheet 9 - Top Customers...")

ws9 = wb.create_sheet("Top Customers")

ws9.merge_cells("A1:E1")
ws9["A1"]           = "Top 15 Customers by Total Sales"
ws9["A1"].font      = title_font
ws9["A1"].alignment = center_align
ws9.row_dimensions[1].height = 30

ws9.append(list(top_cust.columns))
style_header(ws9, 2, len(top_cust.columns))

for i, row in enumerate(top_cust.itertuples(index=False)):
    ws9.append(list(row))
    style_row(ws9, i + 3, len(top_cust.columns), alternate=(i % 2 == 0))

auto_width(ws9)
print("Sheet 9 done!")

print("Creating Sheet 10 - Profit Rate...")

ws10 = wb.create_sheet("Profit Rate")

ws10.merge_cells("A1:E1")
ws10["A1"]           = "Profit Rate by Category"
ws10["A1"].font      = title_font
ws10["A1"].alignment = center_align
ws10.row_dimensions[1].height = 30

ws10.append(list(profit_rate.columns))
style_header(ws10, 2, len(profit_rate.columns))

for i, row in enumerate(profit_rate.itertuples(index=False)):
    ws10.append(list(row))
    row_num  = i + 3
    style_row(ws10, row_num, len(profit_rate.columns), alternate=(i % 2 == 0))
    rate_cell = ws10.cell(row=row_num, column=5)
    try:
        val = float(rate_cell.value)
        if val >= 70:
            rate_cell.fill = green_fill
            rate_cell.font = Font(bold=True, color="375623")
        elif val >= 50:
            rate_cell.fill = gold_fill
        else:
            rate_cell.fill = red_fill
            rate_cell.font = Font(bold=True, color="9C0006")
    except:
        pass

auto_width(ws10)
print("Sheet 10 done!")

print("Creating Sheet 11 - ML Results...")

ws11 = wb.create_sheet("ML Results")

ws11.merge_cells("A1:D1")
ws11["A1"]           = "Machine Learning Model Performance"
ws11["A1"].font      = title_font
ws11["A1"].alignment = center_align
ws11.row_dimensions[1].height = 30

ws11.append(list(ml_results.columns))
style_header(ws11, 2, len(ml_results.columns))

for i, row in enumerate(ml_results.itertuples(index=False)):
    ws11.append(list(row))
    row_num = i + 3
    style_row(ws11, row_num, len(ml_results.columns), alternate=(i % 2 == 0))
    r2_cell = ws11.cell(row=row_num, column=4)
    try:
        val = float(r2_cell.value)
        if val >= 0.9:
            r2_cell.fill = green_fill
            r2_cell.font = Font(bold=True, color="375623")
        elif val >= 0.7:
            r2_cell.fill = gold_fill
        else:
            r2_cell.fill = red_fill
            r2_cell.font = Font(bold=True, color="9C0006")
    except:
        pass

auto_width(ws11)
print("Sheet 11 done!")

wb.save("grocery_sales_report.xlsx")
print("\nSaved: grocery_sales_report.xlsx")
print("Excel report complete!")